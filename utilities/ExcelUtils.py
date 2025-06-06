import os
import openpyxl

def get_excel_file_path(file_name):
    """
    Return the absolute file path to the Excel file inside ExcelFiles folder
    relative to this utilities folder.
    """
    base_dir = os.path.dirname(os.path.dirname(__file__))
    return os.path.join(base_dir, 'ExcelFiles', file_name)


def get_row_count(file_name, sheet_name):
    file_path = get_excel_file_path(file_name)
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]
    return sheet.max_row


def get_column_count(file_name, sheet_name):
    file_path = get_excel_file_path(file_name)
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]
    return sheet.max_column


def get_cell_data(file_name, sheet_name, row_number, column_number):
    file_path = get_excel_file_path(file_name)
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]
    return sheet.cell(row=row_number, column=column_number).value


def set_cell_data(file_name, sheet_name, row_number, column_number, data):
    file_path = get_excel_file_path(file_name)
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]
    sheet.cell(row=row_number, column=column_number).value = data
    workbook.save(file_path)


def get_data_from_excel(file_name, sheet_name):
    file_path = get_excel_file_path(file_name)
    final_list = []
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]
    total_row = sheet.max_row
    total_column = sheet.max_column

    for r in range(2, total_row + 1):
        row_list = []
        for c in range(1, total_column + 1):
            row_list.append(sheet.cell(row=r, column=c).value)
        final_list.append(row_list)

    return final_list
